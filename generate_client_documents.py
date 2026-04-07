from __future__ import annotations

import ast
import json
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
FIGURES_DIR = ROOT / "figures"
DELIVERABLES_DIR = ROOT / "deliverables"
INTERNAL_DELIVERABLES_DIR = ROOT / "archive" / "internal-deliverables"

DELIVERABLES_DIR.mkdir(parents=True, exist_ok=True)
INTERNAL_DELIVERABLES_DIR.mkdir(parents=True, exist_ok=True)

FINAL_CSV_PATH = DATA_DIR / "PSID_Ranked_Questions_Final.csv"
SUMMARY_PATH = DATA_DIR / "psid_artifact_summary.json"

FIG_TOP = FIGURES_DIR / "fig_top_ranked_questions.png"
FIG_TOGGLE = FIGURES_DIR / "fig_toggle_comparison.png"
FIG_TIME = FIGURES_DIR / "fig_time_budget.png"
FIG_UTILITY = FIGURES_DIR / "fig_utility_vs_burden.png"

MASTER_XLSX = INTERNAL_DELIVERABLES_DIR / "Master_Questionnaire.xlsx"
DEPLOYABLE_PDF = INTERNAL_DELIVERABLES_DIR / "Deployable_Questionnaire.pdf"
CODEBOOK_PDF = INTERNAL_DELIVERABLES_DIR / "Codebook.pdf"
REPORT_PDF = INTERNAL_DELIVERABLES_DIR / "Final_Report.pdf"

PAGE_W = 1654
PAGE_H = 2339
MARGIN = 90

BG = (248, 249, 251)
WHITE = (255, 255, 255)
INK = (31, 41, 55)
MUTED = (107, 114, 128)
NAVY = (30, 58, 95)
BLUE = (37, 99, 235)
GREEN = (5, 150, 105)
AMBER = (217, 119, 6)
RED = (220, 38, 38)
LINE = (229, 231, 235)
SOFT_BLUE = (248, 251, 255)
SOFT_AMBER = (255, 249, 241)
SOFT_GREEN = (244, 251, 247)
SOFT_RED = (255, 245, 245)


def load_font(size: int, bold: bool = False):
    candidates = []
    if bold:
        candidates.extend(
            [
                "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
                "/Library/Fonts/Arial Bold.ttf",
                "/System/Library/Fonts/Supplemental/Helvetica.ttc",
            ]
        )
    else:
        candidates.extend(
            [
                "/System/Library/Fonts/Supplemental/Arial.ttf",
                "/Library/Fonts/Arial.ttf",
                "/System/Library/Fonts/Supplemental/Helvetica.ttc",
            ]
        )
    for candidate in candidates:
        if Path(candidate).exists():
            try:
                return ImageFont.truetype(candidate, size=size)
            except OSError:
                continue
    return ImageFont.load_default()


FONT_TITLE = load_font(52, bold=True)
FONT_SUBTITLE = load_font(24)
FONT_H1 = load_font(34, bold=True)
FONT_H2 = load_font(26, bold=True)
FONT_BODY = load_font(22)
FONT_BODY_BOLD = load_font(22, bold=True)
FONT_SMALL = load_font(18)
FONT_KPI = load_font(40, bold=True)
FONT_TABLE = load_font(18)
FONT_TABLE_BOLD = load_font(18, bold=True)
FONT_MONO = load_font(22)


def new_page(title: str, subtitle: str | None = None):
    page = Image.new("RGB", (PAGE_W, PAGE_H), BG)
    draw = ImageDraw.Draw(page)
    draw.text((MARGIN, 48), title, font=FONT_TITLE, fill=NAVY)
    if subtitle:
        draw.text((MARGIN + 2, 118), subtitle, font=FONT_SUBTITLE, fill=MUTED)
    return page, draw


def rounded(draw, box, *, fill=WHITE, outline=LINE, radius=26, width=2):
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def text_height(draw, text: str, font) -> int:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[3] - bbox[1]


def wrap_lines(draw, text: str, font, max_width: int) -> list[str]:
    words = str(text).split()
    if not words:
        return [""]
    lines: list[str] = []
    current = ""
    for word in words:
        trial = word if not current else f"{current} {word}"
        if draw.textbbox((0, 0), trial, font=font)[2] <= max_width:
            current = trial
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def draw_wrapped(draw, x: int, y: int, width: int, text: str, font, *, fill=INK, gap: int = 8):
    for line in wrap_lines(draw, text, font, width):
        draw.text((x, y), line, font=font, fill=fill)
        y += text_height(draw, line, font) + gap
    return y


def draw_bullets(draw, x: int, y: int, width: int, items: list[str], *, font=FONT_BODY, fill=INK, bullet_fill=BLUE, gap: int = 12):
    for item in items:
        lines = wrap_lines(draw, item, font, width - 28)
        draw.ellipse((x, y + 8, x + 8, y + 16), fill=bullet_fill)
        current_y = y
        for line in lines:
            draw.text((x + 20, current_y), line, font=font, fill=fill)
            current_y += text_height(draw, line, font) + 6
        y = current_y + gap
    return y


def draw_footer(draw, text: str):
    bbox = draw.textbbox((0, 0), text, font=FONT_SMALL)
    draw.text((PAGE_W - MARGIN - (bbox[2] - bbox[0]), PAGE_H - 48), text, font=FONT_SMALL, fill=MUTED)


def draw_kpi(draw, x: int, y: int, w: int, h: int, value: str, label: str, accent):
    rounded(draw, (x, y, x + w, y + h), fill=WHITE)
    draw.rectangle((x, y, x + w, y + 14), fill=accent)
    draw.text((x + 24, y + 24), value, font=FONT_KPI, fill=INK)
    draw_wrapped(draw, x + 24, y + 88, w - 48, label, FONT_SMALL, fill=MUTED, gap=4)


def paste_image(page: Image.Image, image_path: Path, box: tuple[int, int, int, int]):
    if not image_path.exists():
        return
    x1, y1, x2, y2 = box
    width = x2 - x1
    height = y2 - y1
    image = Image.open(image_path).convert("RGB")
    image.thumbnail((width, height))
    paste_x = x1 + (width - image.width) // 2
    paste_y = y1 + (height - image.height) // 2
    page.paste(image, (paste_x, paste_y))


def draw_table(draw, x: int, y: int, widths: list[int], headers: list[str], rows: list[list[str]], *, row_height: int = 64, header_fill=SOFT_BLUE, body_fill=WHITE):
    table_width = sum(widths)
    rounded(draw, (x, y, x + table_width, y + row_height * (len(rows) + 1)), fill=body_fill)
    current_x = x
    for width, header in zip(widths, headers):
        draw.rectangle((current_x, y, current_x + width, y + row_height), fill=header_fill, outline=LINE, width=1)
        draw_wrapped(draw, current_x + 12, y + 12, width - 24, header, FONT_TABLE_BOLD, fill=NAVY, gap=2)
        current_x += width

    current_y = y + row_height
    for row in rows:
        current_x = x
        for width, value in zip(widths, row):
            draw.rectangle((current_x, current_y, current_x + width, current_y + row_height), fill=body_fill, outline=LINE, width=1)
            draw_wrapped(draw, current_x + 12, current_y + 10, width - 24, str(value), FONT_TABLE, fill=INK, gap=2)
            current_x += width
        current_y += row_height
    return current_y


def question_type(text: str) -> str:
    lowered = str(text).lower().strip()
    if lowered.startswith("how did") or "select all that apply" in lowered:
        return "Multiple choice / multi-select"
    if lowered.startswith("how often") or lowered.startswith("how severe") or lowered.startswith("how long") or lowered.startswith("how much") or lowered.startswith("overall how would you rate"):
        return "Single choice / ordinal"
    if lowered.startswith("what is your age") or lowered.startswith("how many"):
        return "Numeric / count"
    if lowered.startswith("can i just check"):
        return "Single choice / verification"
    return "Single choice / binary"


def question_universe(row: pd.Series) -> str:
    source = str(row["source"]).lower()
    category = str(row["toggle_category"])
    if category == "Generic":
        return "Ask all respondents"
    if "covid" in source:
        return "Ask when the active crisis is a pandemic or public-health emergency"
    if "katrina" in source:
        return "Ask when the active crisis is a disaster, displacement, or direct exposure event"
    if "shutdown" in source:
        return "Ask when the active crisis is a shutdown, recession, or acute financial disruption"
    return "Ask only when the active specific protocol applies"


def build_questionnaire_frames(df: pd.DataFrame):
    working = df.copy()
    working["Universe"] = working.apply(question_universe, axis=1)
    working["Question Type"] = working["recommended_wording"].apply(question_type)
    working["Selected"] = working["selected_for_module"].map({True: "Yes", False: "No"})
    working["Ri"] = working["Ri"].round(3)
    working["Pi"] = working["Pi"].round(3)
    working["ri_threshold_score"] = working["ri_threshold_score"].round(3)

    master = working[
        [
            "variable_name",
            "toggle_category",
            "historical_toggle_category",
            "Selected",
            "source",
            "Universe",
            "Question Type",
            "question_text",
            "recommended_wording",
            "Ri",
            "ri_threshold_score",
            "Pi",
        ]
    ].rename(
        columns={
            "variable_name": "Variable Name",
            "toggle_category": "Final Category",
            "historical_toggle_category": "Historical Category",
            "source": "Source",
            "question_text": "Historical Question Text",
            "recommended_wording": "Deployable Question Text",
            "ri_threshold_score": "Threshold Score",
        }
    )

    deployable = working[
        (working["selected_for_module"]) & (~working["constructs"].fillna("").astype(str).str.contains("Demographics"))
    ][["variable_name", "Universe", "source", "recommended_wording", "Ri", "Pi"]].rename(
        columns={
            "variable_name": "Variable Name",
            "source": "Source",
            "recommended_wording": "Question Text",
        }
    )
    return master, deployable


def write_master_questionnaire(master_df: pd.DataFrame):
    master_df.to_excel(MASTER_XLSX, index=False)
    workbook = load_workbook(MASTER_XLSX)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in sheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        max_len = min(max(len(value) for value in values) + 2, 48)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_len
        for cell in column_cells[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(MASTER_XLSX)


def parse_literal_list(value) -> list:
    if isinstance(value, list):
        return value
    if not isinstance(value, str) or not value.strip():
        return []
    try:
        parsed = ast.literal_eval(value)
    except (ValueError, SyntaxError):
        return []
    return parsed if isinstance(parsed, list) else []


def build_interaction_summary(df: pd.DataFrame) -> pd.DataFrame:
    constructs = df["constructs"].fillna("").astype(str)
    interaction_summary = (
        df.assign(
            has_demographics=constructs.str.contains("Demographics"),
            has_non_demographic=constructs.apply(
                lambda text: any(
                    term in text
                    for term in [
                        "Economic / Income",
                        "Employment",
                        "Financial Coping",
                        "Housing / Shelter",
                        "Government Aid",
                        "Trauma / Health",
                    ]
                )
            ),
        )
        .groupby(["has_demographics", "has_non_demographic"], as_index=False)
        .agg(rows=("question_text", "count"), avg_ri=("Ri", "mean"), avg_pi=("Pi", "mean"))
    )
    interaction_summary["avg_ri"] = interaction_summary["avg_ri"].round(3)
    interaction_summary["avg_pi"] = interaction_summary["avg_pi"].round(3)
    return interaction_summary


def build_codebook_pdf(df: pd.DataFrame, summary: dict):
    pages: list[Image.Image] = []
    top_generic = df[df["toggle_category"] == "Generic"].sort_values(["Ri", "Pi"], ascending=False).head(3)
    example_row = top_generic.iloc[0] if not top_generic.empty else df.sort_values(["Ri", "Pi"], ascending=False).iloc[0]
    example_keywords = parse_literal_list(example_row.get("tagged_keywords"))
    keyword_weights = [f"{float(item.get('weight', 0.0)):.2f}" for item in example_keywords if isinstance(item, dict)]
    ui_formula = " + ".join(keyword_weights) if keyword_weights else "Derived from tagged keyword weights"
    min_ri = float(df["Ri"].min())
    max_ri = float(df["Ri"].max())
    raw_threshold_hits = int((df["Ri"] >= 0.70).sum())
    interaction_summary = build_interaction_summary(df)
    demographics_only = interaction_summary[
        (interaction_summary["has_demographics"]) & (~interaction_summary["has_non_demographic"])
    ]
    non_demographic = interaction_summary[
        (~interaction_summary["has_demographics"]) & (interaction_summary["has_non_demographic"])
    ]
    demographics_only_row = demographics_only.iloc[0] if not demographics_only.empty else {"rows": 0, "avg_ri": 0.0, "avg_pi": 0.0}
    non_demographic_row = non_demographic.iloc[0] if not non_demographic.empty else {"rows": 0, "avg_ri": 0.0, "avg_pi": 0.0}

    page, draw = new_page("Codebook", "Technical documentation for the refreshed Generic/Specific binary model")
    draw_kpi(draw, MARGIN, 190, 300, 170, str(summary["rows"]), "Ranked questions", BLUE)
    draw_kpi(draw, MARGIN + 330, 190, 300, 170, str(summary["selected_rows"]), "Selected questions", GREEN)
    draw_kpi(draw, MARGIN + 660, 190, 300, 170, f"{summary['selected_minutes']:.2f}", "Selected minutes", AMBER)
    draw_kpi(draw, MARGIN + 990, 190, 300, 170, f"{summary['avg_pi']:.3f}", "Average Pi", RED)
    rounded(draw, (MARGIN, 410, PAGE_W - MARGIN, 930), fill=WHITE)
    draw.text((MARGIN + 30, 450), "Table of contents", font=FONT_H1, fill=NAVY)
    toc_items = [
        "1. Purpose and scope",
        "2. Files and workflow roles",
        "3. End-to-end pipeline design",
        "4. Core functions and what they do",
        "5. Scoring formulas and calculations",
        "6. Binary Generic vs Specific categorization",
        "7. Demographic interaction notes",
        "8. Output schema and deliverables",
    ]
    draw_bullets(draw, MARGIN + 40, 530, PAGE_W - 2 * MARGIN - 80, toc_items, bullet_fill=BLUE)
    rounded(draw, (MARGIN, 1000, PAGE_W - MARGIN, 1880), fill=SOFT_BLUE)
    draw.text((MARGIN + 30, 1040), "What changed in this refresh", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        1130,
        PAGE_W - 2 * MARGIN - 80,
        [
            "The final category model now uses only Generic and Specific.",
            "The previous Financial Crisis label is retained only as a historical traceability field.",
            "The 0.7 threshold is applied to a normalized Ri score so the split is mathematically meaningful.",
            "The final CSV is rebuilt from the raw integrated source during every production run.",
        ],
        bullet_fill=AMBER,
    )
    draw_footer(draw, "Codebook.pdf")
    pages.append(page)

    page, draw = new_page("Workflow Design", "How the raw integrated CSV becomes the final scored client package")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 870), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Pipeline", font=FONT_H1, fill=NAVY)
    code_text = (
        "PSID_Ranked_Questions_Katrina_Integrated.csv\n"
        "  -> normalize_ranked_questions()\n"
        "  -> parse_keywords() / tag_keywords() / extract_constructs()\n"
        "  -> compute_utility() / compute_burden() / Ri\n"
        "  -> assign_binary_categories()\n"
        "  -> compute_augmented_scores()\n"
        "  -> select_for_time_budget()\n"
        "  -> PSID_Ranked_Questions_Final.csv + summary + PDFs"
    )
    rounded(draw, (MARGIN + 30, 300, PAGE_W - MARGIN - 30, 780), fill=SOFT_BLUE)
    draw_wrapped(draw, MARGIN + 60, 340, PAGE_W - 2 * MARGIN - 120, code_text, FONT_MONO, fill=NAVY, gap=10)
    rounded(draw, (MARGIN, 940, PAGE_W - MARGIN, 1960), fill=SOFT_GREEN)
    draw.text((MARGIN + 30, 980), "Core files", font=FONT_H1, fill=NAVY)
    rows = [
        ["PSID_Ranked_Questions_Katrina_Integrated.csv", "Raw integrated source bank"],
        ["PSID_NLP_Crisis_Module_Structure.py", "Reusable helpers and thresholds"],
        ["generate_psid_artifacts.py", "Production rebuild script"],
        ["PSID_Ranked_Questions_Final.csv", "Authoritative scored output"],
        ["PSID_NLP_Crisis_Module_Final.ipynb", "Validation notebook"],
    ]
    draw_table(draw, MARGIN + 30, 1060, [520, 620], ["File", "Role"], rows, row_height=110, header_fill=SOFT_GREEN)
    draw_footer(draw, "Codebook.pdf")
    pages.append(page)

    page, draw = new_page("Functions and Formulas", "Simple descriptions of the functions that matter for review")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 1180), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Key functions", font=FONT_H1, fill=NAVY)
    function_rows = [
        ["build_ranked_dataset()", "Rebuilds the final dataframe from the raw source"],
        ["assign_binary_categories()", "Computes threshold score and Generic/Specific label"],
        ["compute_augmented_scores()", "Adds Pi, redundancy, construct bonus, and portability"],
        ["select_for_time_budget()", "Greedy selection under the 30-minute cap"],
        ["suggest_deployable_wording()", "Converts historical items into deployable wording"],
        ["build_variable_names()", "Creates stable GEN_ and SPC_ identifiers"],
    ]
    draw_table(draw, MARGIN + 30, 300, [420, 720], ["Function", "What it does"], function_rows, row_height=120)
    rounded(draw, (MARGIN, 1260, PAGE_W - MARGIN, 1980), fill=SOFT_AMBER)
    draw.text((MARGIN + 30, 1300), "Core formulas", font=FONT_H1, fill=NAVY)
    formulas = [
        "Ui = sum(keyword weights)",
        "Bi = max(0.10 * word_count + 0.20 * complexity, 0.1)",
        "Ri = Ui / Bi",
        "ri_threshold_score = (Ri - min(Ri)) / (max(Ri) - min(Ri))",
        "Generic if ri_threshold_score >= 0.70 else Specific",
        "Pi = augmented_utility / [Bi * (1 + 0.65 * redundancy_scaled)]",
    ]
    draw_bullets(draw, MARGIN + 40, 1390, PAGE_W - 2 * MARGIN - 80, formulas, bullet_fill=AMBER)
    draw_footer(draw, "Codebook.pdf")
    pages.append(page)

    page, draw = new_page("Worked Example", "Current top Generic item under the refreshed binary model")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 1440), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Calculation table", font=FONT_H1, fill=NAVY)
    example_rows = [
        ["Ui", f"{float(example_row['Ui']):.3f}", ui_formula],
        [
            "Bi",
            f"{float(example_row['Bi']):.3f}",
            f"max(0.10 * {int(example_row['word_count'])} + 0.20 * {float(example_row['complexity']):.1f}, 0.1)",
        ],
        ["Ri", f"{float(example_row['Ri']):.3f}", f"{float(example_row['Ui']):.2f} / {float(example_row['Bi']):.2f}"],
        [
            "ri_threshold_score",
            f"{float(example_row['ri_threshold_score']):.3f}",
            f"({float(example_row['Ri']):.3f} - {min_ri:.3f}) / ({max_ri:.3f} - {min_ri:.3f})",
        ],
        ["Category", str(example_row["toggle_category"]), f"{float(example_row['ri_threshold_score']):.3f} >= 0.70"],
        [
            "Pi",
            f"{float(example_row['Pi']):.3f}",
            "Enhanced priority after IDF, construct, richness, and redundancy adjustments",
        ],
    ]
    draw_table(draw, MARGIN + 30, 300, [260, 180, 820], ["Field", "Value", "Calculation"], example_rows, row_height=120)
    rounded(draw, (MARGIN, 1520, PAGE_W - MARGIN, 1980), fill=SOFT_BLUE)
    draw.text((MARGIN + 30, 1560), "Top Generic rows", font=FONT_H1, fill=NAVY)
    generic_rows = [
        [row.variable_name, row.recommended_wording, f"{row.Ri:.3f}", f"{row.Pi:.3f}"]
        for row in top_generic.itertuples()
    ]
    draw_table(draw, MARGIN + 30, 1640, [360, 640, 120, 120], ["Variable", "Recommended wording", "Ri", "Pi"], generic_rows, row_height=110)
    draw_footer(draw, "Codebook.pdf")
    pages.append(page)

    page, draw = new_page("Threshold and Interaction Notes", "Why the binary split is normalized and how demographics behave")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 860), fill=SOFT_RED)
    draw.text((MARGIN + 30, 230), "Raw-threshold issue", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        320,
        PAGE_W - 2 * MARGIN - 80,
        [
            f"Minimum raw Ri in the refreshed corpus is {min_ri:.3f} and maximum raw Ri is {max_ri:.3f}.",
            f"A direct raw Ri >= 0.7 rule would classify {raw_threshold_hits} of {len(df)} rows as Generic.",
            "The production pipeline therefore applies the 0.7 rule to min-max normalized Ri.",
        ],
        bullet_fill=RED,
    )
    rounded(draw, (MARGIN, 940, PAGE_W - MARGIN, 1980), fill=WHITE)
    draw.text((MARGIN + 30, 980), "Interaction summary", font=FONT_H1, fill=NAVY)
    rows = [
        [
            "Demographics only",
            str(int(demographics_only_row["rows"])),
            f"{float(demographics_only_row['avg_ri']):.3f}",
            f"{float(demographics_only_row['avg_pi']):.3f}",
        ],
        [
            "Non-demographic crisis items",
            str(int(non_demographic_row["rows"])),
            f"{float(non_demographic_row['avg_ri']):.3f}",
            f"{float(non_demographic_row['avg_pi']):.3f}",
        ],
    ]
    draw_table(draw, MARGIN + 30, 1100, [420, 160, 160, 160], ["Group", "Rows", "Average Ri", "Average Pi"], rows, row_height=120)
    draw_wrapped(
        draw,
        MARGIN + 30,
        1460,
        PAGE_W - 2 * MARGIN - 60,
        "Interpretation: demographic items remain useful for traceability, but they carry less direct crisis signal than housing, trauma, aid, employment, or economic-loss items. That is why the deployable questionnaire excludes demographics even though the master questionnaire retains them.",
        FONT_BODY,
        fill=INK,
    )
    draw_footer(draw, "Codebook.pdf")
    pages.append(page)

    page, draw = new_page("Outputs and Benchmarks", "Current client package outputs and refreshed production metrics")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 900), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Deliverables", font=FONT_H1, fill=NAVY)
    rows = [
        ["Master_Questionnaire.xlsx", "Internal review spreadsheet with all 52 rows"],
        ["Deployable_Questionnaire.pdf", "Selected deployable questionnaire excluding demographics"],
        ["Codebook.pdf", "Technical codebook"],
        ["Final_Report.pdf", "Stakeholder-facing report"],
    ]
    draw_table(draw, MARGIN + 30, 320, [420, 720], ["Deliverable", "Purpose"], rows, row_height=120)
    rounded(draw, (MARGIN, 980, PAGE_W - MARGIN, 1980), fill=SOFT_GREEN)
    draw.text((MARGIN + 30, 1020), "Benchmark summary", font=FONT_H1, fill=NAVY)
    benchmark_rows = [
        ["Rows", str(summary["rows"])],
        ["Selected rows", str(summary["selected_rows"])],
        ["Selected minutes", f"{summary['selected_minutes']:.2f}"],
        ["Average Ri", f"{summary['avg_ri']:.3f}"],
        ["Average Pi", f"{summary['avg_pi']:.3f}"],
        ["Generic rows", str(summary["toggle_counts"]["Generic"])],
        ["Specific rows", str(summary["toggle_counts"]["Specific"])],
    ]
    draw_table(draw, MARGIN + 30, 1140, [360, 220], ["Metric", "Value"], benchmark_rows, row_height=100, header_fill=SOFT_GREEN)
    paste_image(page, FIG_TIME, (900, 1120, PAGE_W - MARGIN - 30, 1880))
    draw_footer(draw, "Codebook.pdf")
    pages.append(page)

    pages[0].save(CODEBOOK_PDF, save_all=True, append_images=pages[1:])


def build_report_pdf(df: pd.DataFrame, summary: dict):
    pages: list[Image.Image] = []
    generic_df = df[df["toggle_category"] == "Generic"].sort_values(["Ri", "Pi"], ascending=False)
    specific_df = df[df["toggle_category"] == "Specific"].sort_values(["Ri", "Pi"], ascending=False)
    selected_by_source = pd.DataFrame(summary["selected_source_counts"].items(), columns=["Source", "Selected"])
    interaction_summary = build_interaction_summary(df)
    demographics_only = interaction_summary[
        (interaction_summary["has_demographics"]) & (~interaction_summary["has_non_demographic"])
    ]
    non_demographic = interaction_summary[
        (~interaction_summary["has_demographics"]) & (interaction_summary["has_non_demographic"])
    ]
    demographics_only_row = demographics_only.iloc[0] if not demographics_only.empty else {"rows": 0, "avg_ri": 0.0, "avg_pi": 0.0}
    non_demographic_row = non_demographic.iloc[0] if not non_demographic.empty else {"rows": 0, "avg_ri": 0.0, "avg_pi": 0.0}

    page, draw = new_page("Final Report", "Stakeholder-facing summary of the refreshed PSID crisis-module package")
    draw_kpi(draw, MARGIN, 190, 300, 170, str(summary["rows"]), "Ranked questions", BLUE)
    draw_kpi(draw, MARGIN + 330, 190, 300, 170, str(summary["selected_rows"]), "Selected questions", GREEN)
    draw_kpi(draw, MARGIN + 660, 190, 300, 170, f"{summary['selected_minutes']:.2f}", "Selected minutes", AMBER)
    draw_kpi(draw, MARGIN + 990, 190, 300, 170, f"{summary['avg_ri']:.3f}", "Average Ri", RED)
    rounded(draw, (MARGIN, 430, PAGE_W - MARGIN, 960), fill=WHITE)
    draw.text((MARGIN + 30, 470), "Executive summary", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        560,
        PAGE_W - 2 * MARGIN - 80,
        [
            "The final client-facing category model is now Generic versus Specific.",
            "The previous Financial Crisis label has been removed as a final category and retained only as a traceability field.",
            "The final selected module contains 28 questions and stays under the 30-minute administration cap.",
            "The package now separates technical explanation, deployable wording, and stakeholder reporting into distinct deliverables.",
        ],
        bullet_fill=BLUE,
    )
    rounded(draw, (MARGIN, 1040, PAGE_W - MARGIN, 1880), fill=SOFT_BLUE)
    draw.text((MARGIN + 30, 1080), "Table of contents", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        1180,
        PAGE_W - 2 * MARGIN - 80,
        [
            "1. Background and objective",
            "2. Methodology",
            "3. Results",
            "4. Discussion of skew",
            "5. Deliverables summary",
            "6. Conclusion and next steps",
        ],
        bullet_fill=AMBER,
    )
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("Background and Objective", "Why the package was refreshed and what changed")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 1980), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Project goal", font=FONT_H1, fill=NAVY)
    y = draw_wrapped(
        draw,
        MARGIN + 30,
        320,
        PAGE_W - 2 * MARGIN - 60,
        "The PSID archive contains crisis-related items from multiple historical contexts. The client requirement was to turn that archive into a simpler and more defensible package: a cleaner deployable questionnaire, a technical codebook, and a stakeholder-facing report that explains the shortlist without overemphasizing finance-only content.",
        FONT_BODY,
        fill=INK,
    )
    y += 30
    draw.text((MARGIN + 30, y), "What changed", font=FONT_H2, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        y + 70,
        PAGE_W - 2 * MARGIN - 80,
        [
            "Binary Generic/Specific category model replaces the older multi-label final structure.",
            "Financial Crisis is no longer treated as a separate final module category.",
            "The 0.7 client threshold is implemented on normalized Ri rather than raw Ri.",
            "The final CSV is rebuilt from the raw integrated source during every refresh.",
        ],
        bullet_fill=GREEN,
    )
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("Methodology", "How the ranking and categorization logic works in plain language")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 920), fill=SOFT_BLUE)
    draw.text((MARGIN + 30, 230), "Pipeline overview", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        320,
        PAGE_W - 2 * MARGIN - 80,
        [
            "Load 52 historical questions from the integrated source file.",
            "Reuse stored keyword lists and map them to constructs such as Employment, Government Aid, and Trauma / Health.",
            "Calculate utility, burden, and the baseline Ri ranking score.",
            "Normalize Ri and apply the 0.7 threshold to assign Generic or Specific.",
            "Compute the enhanced Pi score using rarity, construct coverage, portability, and redundancy penalties.",
            "Select the final module greedily under the 30-minute cap.",
        ],
        bullet_fill=BLUE,
    )
    rounded(draw, (MARGIN, 1000, PAGE_W - MARGIN, 1980), fill=WHITE)
    draw.text((MARGIN + 30, 1040), "Key formulas", font=FONT_H1, fill=NAVY)
    formulas = [
        ["Ri", "Ui / Bi"],
        ["Bi", "max(0.10 * word_count + 0.20 * complexity, 0.1)"],
        ["minutes", "(word_count * 7) / 60"],
        ["ri_threshold_score", "(Ri - min(Ri)) / (max(Ri) - min(Ri))"],
        ["Pi", "augmented_utility / [Bi * (1 + 0.65 * redundancy_scaled)]"],
    ]
    draw_table(draw, MARGIN + 30, 1160, [260, 860], ["Field", "Formula"], formulas, row_height=110)
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("Results", "Updated counts, source mix, and top questions")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 960), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Headline metrics", font=FONT_H1, fill=NAVY)
    metrics_rows = [
        ["Selected questions", str(summary["selected_rows"])],
        ["Selected minutes", f"{summary['selected_minutes']:.2f}"],
        ["Average Ri", f"{summary['avg_ri']:.3f}"],
        ["Average Pi", f"{summary['avg_pi']:.3f}"],
        ["Generic rows", str(summary["toggle_counts"]["Generic"])],
        ["Specific rows", str(summary["toggle_counts"]["Specific"])],
    ]
    draw_table(draw, MARGIN + 30, 320, [320, 180], ["Metric", "Value"], metrics_rows, row_height=100)
    draw_table(
        draw,
        650,
        320,
        [360, 180],
        ["Source", "Selected"],
        selected_by_source.values.tolist(),
        row_height=100,
    )
    rounded(draw, (MARGIN, 1040, PAGE_W - MARGIN, 1980), fill=SOFT_GREEN)
    draw.text((MARGIN + 30, 1080), "Top Generic and Specific examples", font=FONT_H1, fill=NAVY)
    generic_rows = [[row.variable_name, row.recommended_wording, f"{row.Ri:.3f}"] for row in generic_df.head(3).itertuples()]
    specific_rows = [[row.variable_name, row.recommended_wording, f"{row.Ri:.3f}"] for row in specific_df.head(5).itertuples()]
    draw.text((MARGIN + 30, 1180), "Generic", font=FONT_H2, fill=NAVY)
    draw_table(draw, MARGIN + 30, 1240, [360, 560, 120], ["Variable", "Question", "Ri"], generic_rows, row_height=110, header_fill=SOFT_GREEN)
    draw.text((MARGIN + 30, 1640), "Specific", font=FONT_H2, fill=NAVY)
    draw_table(draw, MARGIN + 30, 1700, [360, 560, 120], ["Variable", "Question", "Ri"], specific_rows, row_height=95, header_fill=SOFT_GREEN)
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("Analytical Figures", "Selected figures from the refreshed artifact bundle")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 1050), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Top-ranked questions", font=FONT_H1, fill=NAVY)
    paste_image(page, FIG_TOP, (MARGIN + 30, 310, PAGE_W - MARGIN - 30, 980))
    rounded(draw, (MARGIN, 1120, PAGE_W - MARGIN, 1980), fill=WHITE)
    draw.text((MARGIN + 30, 1160), "Time budget by category", font=FONT_H1, fill=NAVY)
    paste_image(page, FIG_TIME, (MARGIN + 30, 1240, PAGE_W - MARGIN - 30, 1910))
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("Discussion of Skew", "Why financial questions appear fewer and why demographics fall away")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 1980), fill=SOFT_AMBER)
    draw.text((MARGIN + 30, 230), "Main interpretation", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        320,
        PAGE_W - 2 * MARGIN - 80,
        [
            "The earlier output looked skewed partly because Financial Crisis appeared as its own visible module label.",
            "In the refreshed package, financial items compete inside Specific alongside disaster and pandemic items.",
            "Several financial items are close variants of one another, so redundancy penalties push some of them down.",
            "Katrina disaster items remain numerous and competitive, which reduces the share of finance-labeled items in the final shortlist.",
            "Demographic-only rows score lower on average because they add context rather than direct crisis signal.",
        ],
        bullet_fill=AMBER,
    )
    draw.text((MARGIN + 30, 970), "Demographic versus non-demographic averages", font=FONT_H2, fill=NAVY)
    rows = [
        [
            "Demographics only",
            str(int(demographics_only_row["rows"])),
            f"{float(demographics_only_row['avg_ri']):.3f}",
            f"{float(demographics_only_row['avg_pi']):.3f}",
        ],
        [
            "Non-demographic crisis items",
            str(int(non_demographic_row["rows"])),
            f"{float(non_demographic_row['avg_ri']):.3f}",
            f"{float(non_demographic_row['avg_pi']):.3f}",
        ],
    ]
    draw_table(draw, MARGIN + 30, 1040, [420, 160, 160, 160], ["Group", "Rows", "Average Ri", "Average Pi"], rows, row_height=110, header_fill=SOFT_AMBER)
    draw_wrapped(
        draw,
        MARGIN + 30,
        1540,
        PAGE_W - 2 * MARGIN - 60,
        "Takeaway: the refreshed package still retains high-value financial material, but it no longer presents finance as if it were the only or dominant crisis dimension. The selected module is better read as a broader shock-and-response instrument covering economic disruption, aid exposure, housing instability, and trauma-related impact.",
        FONT_BODY,
        fill=INK,
    )
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("Deliverables Summary", "What is included for client use")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 980), fill=WHITE)
    draw.text((MARGIN + 30, 230), "Package contents", font=FONT_H1, fill=NAVY)
    rows = [
        ["Master_Questionnaire.xlsx", "Internal spreadsheet with all rows, categories, source, universe, and scores"],
        ["Deployable_Questionnaire.pdf", "Questionnaire for deployment using cleaned wording and no demographics"],
        ["Codebook.pdf", "Technical documentation of functions, formulas, schema, and thresholds"],
        ["Final_Report.pdf", "Stakeholder-facing explanation of results and interpretation"],
    ]
    draw_table(draw, MARGIN + 30, 320, [420, 720], ["Deliverable", "Use"], rows, row_height=120)
    rounded(draw, (MARGIN, 1060, PAGE_W - MARGIN, 1980), fill=SOFT_BLUE)
    draw.text((MARGIN + 30, 1100), "Recommended next steps", font=FONT_H1, fill=NAVY)
    draw_bullets(
        draw,
        MARGIN + 40,
        1190,
        PAGE_W - 2 * MARGIN - 80,
        [
            "Review whether one of the current Generic items should be manually demoted if the client wants a stricter event-agnostic core.",
            "Compare the deployable questionnaire row by row against Tom's preferred formatting template before sign-off.",
            "Use the dashboard and figures as supplemental analytical materials rather than required client-facing content unless requested.",
        ],
        bullet_fill=BLUE,
    )
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    page, draw = new_page("References", "Selected sources used to support interpretation and reporting")
    rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 1980), fill=WHITE)
    refs = [
        "Panel Study of Income Dynamics (PSID): Main Interview, 2021. ICPSR 39190.",
        "Fifty Years of the Panel Study of Income Dynamics: Past, Present, and Future.",
        "User Guide for the 2019 Interviewing Year.",
        "User Guide for the 2021 Interviewing Year.",
        "User Guide for the 2023 Interviewing Year.",
        "PSID COVID-19 Measures documentation.",
        "Economic Impact Payments and Household Spending During the Pandemic. NBER.",
        "The Economic Impacts of COVID-19: Evidence from a New Public Database Built Using Private Sector Data.",
        "Internal analytical source: PSID Data Modeling Report Prompt Generation.md.",
    ]
    draw.text((MARGIN + 30, 230), "References", font=FONT_H1, fill=NAVY)
    draw_bullets(draw, MARGIN + 40, 320, PAGE_W - 2 * MARGIN - 80, refs, bullet_fill=GREEN)
    draw_footer(draw, "Final_Report.pdf")
    pages.append(page)

    pages[0].save(REPORT_PDF, save_all=True, append_images=pages[1:])


def build_deployable_pdf(deployable_df: pd.DataFrame):
    rows_per_page = 7
    pages: list[Image.Image] = []
    total_pages = math.ceil(len(deployable_df) / rows_per_page)
    for page_number in range(total_pages):
        page, draw = new_page(
            "Deployable Questionnaire",
            "Tom template: Variable Name | Universe | Source | Question Text",
        )
        rounded(draw, (MARGIN, 190, PAGE_W - MARGIN, 2140), fill=WHITE)
        slice_df = deployable_df.iloc[page_number * rows_per_page : (page_number + 1) * rows_per_page]
        rows = [
            [row["Variable Name"], row["Universe"], row["Source"], row["Question Text"]]
            for _, row in slice_df.iterrows()
        ]
        draw_table(draw, MARGIN + 30, 260, [280, 420, 220, 520], ["Variable Name", "Universe", "Source", "Question Text"], rows, row_height=220)
        draw_footer(draw, f"Deployable_Questionnaire.pdf | Page {page_number + 1} of {total_pages}")
        pages.append(page)
    pages[0].save(DEPLOYABLE_PDF, save_all=True, append_images=pages[1:])


def main():
    df = pd.read_csv(FINAL_CSV_PATH)
    summary = json.loads(SUMMARY_PATH.read_text())
    master_df, deployable_df = build_questionnaire_frames(df)
    write_master_questionnaire(master_df)
    build_deployable_pdf(deployable_df)
    build_codebook_pdf(df, summary)
    build_report_pdf(df, summary)
    print(json.dumps(
        {
            "master_questionnaire": str(MASTER_XLSX.relative_to(ROOT)),
            "deployable_questionnaire": str(DEPLOYABLE_PDF.relative_to(ROOT)),
            "codebook_pdf": str(CODEBOOK_PDF.relative_to(ROOT)),
            "report_pdf": str(REPORT_PDF.relative_to(ROOT)),
            "deployable_rows": int(len(deployable_df)),
        },
        indent=2,
    ))


if __name__ == "__main__":
    main()
